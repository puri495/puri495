# programme for ATM transactions
# openpyxl is python library to operate with excel sheets
import openpyxl as xl
wb = xl.load_workbook('Atm_project_data1.xlsx')
sheet = wb['Sheet1']
# load the excel data into list


def ret_list(cl):
    lists = []
    for i in range(2,sheet.max_row+1):
        cell = sheet.cell(i, cl)
        lists.append(cell.value)
    return lists


users = ret_list(cl=1)
pins = ret_list(cl=2)
amounts = ret_list(cl=3)
print('*****WELCOME*****\n')
first = int(input('''
        1.CREATE ACCOUNT:
        2.LOGIN         :\n
        '''))
count = 1
if first == 1:
    r = sheet.max_row + 1
    name = input('Enter Your Name:\n')
    name_cell = sheet.cell(r, 1)
    name_cell.value = name
    pin = int(input('Enter Your 4 digit PIN:\n'))
    pin_cell = sheet.cell(r, 2)
    pin_cell.value = pin
    initial_dept = int(input('Enter initial amount gonna Deposit:\n'))
    initial_dept_cell = sheet.cell(r, 3)
    initial_dept_cell.value = initial_dept
    print('COME BACK FOR NEW TRANSACTIONS\n')
    wb.save('Atm_project_data1.xlsx')
    print()
elif first == 2:
    print('Enter User Name:\n')
    user = input()
    if user in users:
        n = users.index(user)
        while count < 4:
            print('ENTER PIN NUM:\n')
            pin = int(input())
            if pin == pins[n]:
                print(''' HOW CAN I ASSIST YOU
                        1.CHECK BALANCE
                        2.WITHDRAW
                        3.DEPOSIT
                        4.PRINT STATEMENT\n''')
                p = int(input())
                if p == 1:
                    print(amounts[n])
                    break
                elif p == 2:
                    k = int(input('Enter the amount to withdraw\n'))
                    a = int(amounts[n])
                    r = a-k
                    amounts.insert(n,r)
                    sheet.cell(row=n+2, column=3, value=r)
                    wb.save('Atm_project_data1.xlsx')
                    print('REMAINING BALANCE IS: ')
                    print(amounts[n])
                    break
                elif p == 3:
                    k = int(input('Enter the amount to deposit\n'))
                    a = int(amounts[n])
                    r = a + k
                    amounts.insert(n, r)
                    sheet.cell(row=n+2, column=3, value=r)
                    wb.save('Atm_project_data1.xlsx')
                    print('AVAILABLE BALANCE IS: \n')
                    print(amounts[n])
                    break
                elif p == 4:
                    print('AVAILABLE BALANCE IS: ')
                    print(amounts[n])
                    break
                else:
                    print('******Choose Correct Option******')
                    count += 1
                    if count > 3:
                        break
            else:
                print('*****************')
                print('Enter Correct PIN')
                print('*****************')
                count = count+1
        if count > 3:
            print('TOO MANY TRIES')
            print(users[n]+' *******YOUR ACCOUNT IS BLOCKED**********')
    else:
        print('*****************')
        print('INVALID USERNAME')
        print('*****************')

else:
    print('Enter Valid Number:')
