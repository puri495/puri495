from flask import Flask, render_template, request, redirect, url_for
import openpyxl as xl

app = Flask(__name__)
wb = xl.load_workbook('Atm_project_data1.xlsx')
sheet = wb['Sheet1']

# Function to load the excel data into list
def ret_list(cl):
    lists = []
    for i in range(2, sheet.max_row + 1):
        cell = sheet.cell(i, cl)
        lists.append(cell.value)
    return lists

users = ret_list(cl=1)
pins = ret_list(cl=2)
amounts = ret_list(cl=3)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/create_account', methods=['GET', 'POST'])
def create_account():
    if request.method == 'POST':
        r = sheet.max_row + 1
        name = request.form['name']
        pin = int(request.form['pin'])
        initial_dept = int(request.form['initial_dept'])

        name_cell = sheet.cell(r, 1)
        name_cell.value = name
        pin_cell = sheet.cell(r, 2)
        pin_cell.value = pin
        initial_dept_cell = sheet.cell(r, 3)
        initial_dept_cell.value = initial_dept

        wb.save('Atm_project_data1.xlsx')
        return redirect(url_for('index'))
    return render_template('create_account.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = request.form['username']
        pin = int(request.form['pin'])

        if user in users:
            n = users.index(user)
            if pin == pins[n]:
                return redirect(url_for('dashboard', user=user))
            else:
                return "Incorrect PIN. Please try again."
        else:
            return "Invalid Username. Please try again."
    return render_template('login.html')

@app.route('/dashboard/<user>')
def dashboard(user):
    n = users.index(user)
    balance = amounts[n]
    return render_template('dashboard.html', user=user, balance=balance)

@app.route('/withdraw', methods=['POST'])
def withdraw():
    user = request.form['user']
    n = users.index(user)
    k = int(request.form['amount'])
    a = int(amounts[n])
    r = a - k
    amounts[n] = r
    sheet.cell(row=n + 2, column=3, value=r)
    wb.save('Atm_project_data1.xlsx')
    return redirect(url_for('dashboard', user=user))

@app.route('/deposit', methods=['POST'])
def deposit():
    user = request.form['user']
    n = users.index(user)
    k = int(request.form['amount'])
    a = int(amounts[n])
    r = a + k
    amounts[n] = r
    sheet.cell(row=n + 2, column=3, value=r)
    wb.save('Atm_project_data1.xlsx')
    return redirect(url_for('dashboard', user=user))

if __name__ == '__main__':
    app.run(debug=True)
